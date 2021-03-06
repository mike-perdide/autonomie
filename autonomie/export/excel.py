# -*- coding: utf-8 -*-
# * Copyright (C) 2012-2013 Croissance Commune
# * Authors:
#       * Arezki Feth <f.a@majerti.fr>;
#       * Miotte Julien <j.m@majerti.fr>;
#       * Pettier Gabriel;
#       * TJEBBES Gaston <g.t@majerti.fr>
#
# This file is part of Autonomie : Progiciel de gestion de CAE.
#
#    Autonomie is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    Autonomie is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with Autonomie.  If not, see <http://www.gnu.org/licenses/>.
#

"""
    Data export module
"""

import openpyxl
from openpyxl.style import Color, Fill
import cStringIO as StringIO
from string import ascii_uppercase

from autonomie.export.utils import write_file_to_request
from autonomie.compute.math_utils import integer_to_amount
from autonomie.models.treasury import ExpenseType
from autonomie.models.treasury import ExpenseKmType
from autonomie.models.treasury import ExpenseTelType


LETTERS = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I',)

Color.LightCyan = "FFE0FFFF"
Color.LightCoral = "FFF08080"
Color.LightGreen = "FF90EE90"
Color.Crimson = "FFDC143C"
Color.header = "FFD9EDF7"
Color.footer = "FFFCF8E3"

def get_xls_column_keys():
    res = []
    for let in ascii_uppercase:
        res.append(let)

    for leta in ascii_uppercase:
        for letb in ascii_uppercase:
            res.append('%s%s'% (leta, letb))
    return res

ASCII_UPPERCASE = get_xls_column_keys()

class ExcelExpense(object):
    """
        Wrapper for excel export of an expense object
    """
    expensekm_columns = [
    {'label':u'Date', 'key':'date', 'letter': 'A'},
    {'label':u'Type de véhicule', 'key':'vehicle', 'letter':'B',
        'last_letter':'C'},
    {'label':u'Lieu de départ', 'key':'start', 'letter':'D',
        'last_letter':'E'},
    {'label':u"Lieu d'arrivée", 'key':'end', 'letter':'F',
        'last_letter':'G'},
    {'label':u'Description/Mission', 'key':'description', 'letter':'H',
        'last_letter':'J'},
    {'label':'Nombre de kms', 'key':'km', 'formatter':integer_to_amount,
        'letter':'K'},
    {'label':'Indemnités', 'key':'total', 'formatter':integer_to_amount,
        'letter':'L', 'number_format':'0.00'}]

    def __init__(self, expensesheet):
        self.book = openpyxl.workbook.Workbook()
        self.model = expensesheet
        self.internal_columns = self.get_columns()
        self.activity_columns = self.get_columns(internal=False)
        self.index = 2

    def get_tel_column(self):
        """
        """
        teltype = ExpenseTelType.query().first()
        if teltype:
            return [{'label':"Téléphonie", 'code':teltype.code}]
        else:
            return []

    def get_km_column(self):
        """
        """
        kmtype = ExpenseKmType.query().first()
        if kmtype:
            return [{'label':"Frais de déplacement", 'code':kmtype.code}]
        else:
            return []

    def get_disabled_types_columns(self):
        """
        """
        types = []
        for line in self.model.lines:
            type_ = line.type_object
            if not type_.active and type_.type != 'expensetel':
                if type_.id not in types:
                    types.append(type_.id)
                    yield {
                         'label':"%s (ce type de frais n'existe plus)"\
                                 % (type_.label),
                         'code':type_.code
                     }

    def get_columns(self, internal=True):
        """
            Retrieve all columns and define a global column attribute
            :param internal: are we asking columns for internal expenses
        """
        columns = []
        # Add the two first columns
        columns.append({'label':'Date', 'key':'date'})
        columns.append({
            'label':'Description',
            'key':'description',
            'additional_cell_nb':3})
        # Telephonic fees are only available as internal expenses
        if internal:
            columns.extend(self.get_tel_column())
        km_columns = self.get_km_column()
        columns.extend(km_columns)

        if len(km_columns) > 0:
            kmtype_code = km_columns[0]['code']
        else:
            kmtype_code = None

        commontypes = ExpenseType.query()\
                .filter(ExpenseType.type=="expense")\
                .filter(ExpenseType.active==True)
        for type_ in commontypes:
            # Here's a hack to allow to group km fee types and displacement fees
            if kmtype_code is not None and type_.code != kmtype_code:
                columns.append({'label':type_.label, 'code':type_.code})

        columns.extend(self.get_disabled_types_columns())

        # Add the last columns
        columns.append({'label':'Tva', 'key':'tva',
            'formatter':integer_to_amount})
        columns.append({'label':'Total', 'key':'total',
            'formatter':integer_to_amount, 'number_format':'0.00'})

        # We set the appropriate letter to each column
        index = 0
        for col in columns:
            letter = ASCII_UPPERCASE[index]
            additional_cell_nb = col.get('additional_cell_nb')
            if additional_cell_nb:
                last_letter = ASCII_UPPERCASE[index + additional_cell_nb]
                index += additional_cell_nb + 1
            else:
                last_letter = letter
                index += 1
            col['letter'] = letter
            col['last_letter'] = last_letter
        return columns

    def get_merged_cells(self, start, end):
        """
            returned merged cells of the current line index
        """
        cell_gap = '{1}{0}:{2}{0}'.format(self.index, start, end)
        self.worksheet.merge_cells(cell_gap)
        cell = self.worksheet.cell('{1}{0}'.format(self.index, start))
        return cell

    def set_color(self, cell, color):
        cell.style.fill.fill_type = Fill.FILL_SOLID
        cell.style.fill.start_color.index = color

    def write_code(self):
        """
            write the company code in the header
        """
        code = self.model.company.code_compta
        if not code:
            code = u"Code non renseigné"
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Code analytique de l'entreprise"
        cell = self.get_merged_cells('E', 'J')
        cell.value = code
        self.index += 1

    def write_user(self):
        """
            write the username in the header
        """
        user = self.model.user
        title = u"%s %s" % (user.lastname, user.firstname)
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Nom de l'entrepreneur"
        cell = self.get_merged_cells('E', 'J')
        cell.value = title
        self.index += 1

    def write_period(self):
        """
            write the period in the header
        """
        period = "01/{0}/{1}".format(self.model.month,
                self.model.year)
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Période de la demande"
        cell = self.get_merged_cells('E', 'J')
        cell.value = period
        self.index += 2

    def get_column_cell(self, column):
        """
            Return the cell corresponding to a given column
        """
        letter = column['letter']
        last_letter = column.get('last_letter', letter)
        return self.get_merged_cells(letter, last_letter)

    def write_table_header(self, columns):
        """
            Write the table's header and its subheader
        """
        for column in columns:
            cell = self.get_column_cell(column)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = Color.header
            cell.style.font.bold = True
            cell.value = column['label']
        self.index += 1
        for column in columns:
            cell = self.get_column_cell(column)
            cell.style.font.bold = True
            cell.value = column.get('code', '')
        self.index += 1

    def get_cell_val(self, line, column):
        """
            return the value for a given cell for the current line
        """
        if line.type_object is None:
            val = ""
        elif column.has_key('key'):
            val = getattr(line, column['key'], '')
            formatter = column.get('formatter')
            if val and formatter:
                val = formatter(val)
        elif column['code'] == line.type_object.code:
            if hasattr(line, 'ht'):
                val = integer_to_amount(line.ht)
            else:
                val = integer_to_amount(line.total)
        else:
            val = ""
        return val

    def set_col_width(self, col_letter, width, force=False):
        """
            Set the width of a given column
        """
        col_dim = self.worksheet.column_dimensions.get(col_letter)
        if col_dim:
            if col_dim.width == -1 or force:
                if width == 0:
                    col_dim.visible=False
                else:
                    col_dim.width = width

    def write_table(self, columns, lines):
        """
            write a table with headers and content
            :param columns: list of dict
            :params lines: list of models to be written
        """
        self.write_table_header(columns)
        for line in lines:
            for column in columns:
                cell = self.get_column_cell(column)
                cell.value = self.get_cell_val(line, column)
                if column.has_key('number_format'):
                    cell.style.number_format.format_code = \
                            column['number_format']
            self.index += 1

        for column in columns:
            cell = self.get_column_cell(column)
            cell.style.font.bold = True
            self.set_color(cell, Color.footer)
            cell.style.number_format.format_code = '0.00'
            if column.has_key('code'):
                val = sum([line.ht for line in lines \
                                     if line.type_object \
                                     and line.type_object.code==column['code']])
                cell.value = integer_to_amount(val)
                if val == 0:
                    self.set_col_width(column['letter'], 0)
                else:
                    self.set_col_width(column['letter'], 13)
            elif column.get('key') == 'description':
                cell.value = "Totaux"
            elif column.get('key') == 'tva':
                cell.value = integer_to_amount(
                        sum([getattr(line, 'tva', 0) for line in lines]))
                cell.style.number_format.format_code = '0.00'
            elif column.get('key') == 'total':
                cell.value = integer_to_amount(
                        sum([line.total for line in lines]))
                cell.style.number_format.format_code = '0.00'
        self.index += 4

    def write_expense_table(self, category, columns):
        """
            write expenses tables for the given category
        """
        lines = [line for line in self.model.lines
                            if line.category == category]
        kmlines = [lin for lin in self.model.kmlines
                            if lin.category == category]
        lines.extend(kmlines)
        self.write_table(columns, lines)
        self.index += 2

    def write_full_line(self, txt, start="A", end="J"):
        """
            Write a full line, merging cells
        """
        cell = self.get_merged_cells(start, end)
        cell.value = txt
        self.index += 1
        return cell

    def write_internal_expenses(self):
        """
            write the internal expense table to the current worksheet
        """
        txt = u"FRAIS DIRECT DE FONCTIONNEMENT (< à 30% DU SALAIRE BRUT \
PAR MOIS)"
        cell = self.write_full_line(txt)
        cell.style.font.color.index = Color.Crimson
        self.write_expense_table('1', self.internal_columns)

    def write_activity_expenses(self):
        """

            write the activity expense table to the current worksheet
        """
        txt = u"FRAIS CONCERNANT DIRECTEMENT VOTRE L'ACTIVITE AUPRES DE VOS \
CLIENTS"
        cell = self.write_full_line(txt)
        cell.style.font.color.index = Color.Crimson
        self.write_expense_table('2', self.activity_columns)

    def write_total(self):
        """
            write the final total
        """
        cell = self.get_merged_cells('A', 'D')
        cell.value = u"Total des frais professionnel à payer"
        cell.style.font.bold = True
        cell.style.font.size = 16
        self.set_color(cell, Color.footer)
        cell = self.get_merged_cells('E', 'E')
        cell.style.font.bold = True
        cell.style.font.size = 16
        cell.value = integer_to_amount(self.model.total)
        cell.style.number_format.format_code = '0.00'
        self.set_color(cell, Color.footer)
        self.index += 2

    def write_accord(self):
        """
            Write the endline
        """
        cell = self.get_merged_cells('B', 'E')
        cell.value = u"Accord après vérification"
        self.index += 1
        self.worksheet.merge_cells(
                start_row=self.index,
                end_row=self.index +4,
                start_column=1,
                end_column=4)

    def write_km_book(self):
        """
            Write the km book associated to this expenses
        """
        self.index = 3
        user = self.model.user
        title = u"Tableau de bord kilométrique de {0} {1}".\
                format(user.lastname, user.firstname)
        cell = self.write_full_line(title)
        cell.style.font.bold = True
        cell.style.font.size = 24
        # index has already been increased
        row_dim = self.worksheet.row_dimensions.get(self.index -1 )
        row_dim.height = 30
        self.index += 2

        self.write_table(self.expensekm_columns, self.model.kmlines)

    def render(self):
        """
            Return the current excel export as a String buffer (StringIO)
        """
        self.worksheet = self.book.worksheets[0]
        self.worksheet.title = "NDF"
        cell = self.write_full_line(u"Feuille de notes de frais")
        cell.style.font.bold = True
        cell.style.font.size = 24
        # index has already been increased
        row_dim = self.worksheet.row_dimensions.get(self.index -1 )
        row_dim.height = 30
        self.index += 2

        self.write_code()
        self.write_user()
        self.write_period()
        self.write_internal_expenses()
        self.write_activity_expenses()
        self.write_total()
        self.write_accord()

        # We set a width to all columns that have no width set (-1)
        for let in ASCII_UPPERCASE:
            self.set_col_width(let, 13)

        self.worksheet = self.book.create_sheet()
        self.worksheet.title = u"Journal de bord"
        self.write_km_book()

        for let in ASCII_UPPERCASE:
            col_dim = self.worksheet.column_dimensions.get(let)
            if col_dim:
                col_dim.width = 13

        result = StringIO.StringIO()
        self.book.save(result)
        return result


def make_excel_view(filename_builder, factory):
    """
        Build an excel view of a model
        :param filename_builder: a callable that take the request as arg and
            return a filename
        :param factory: the Excel factory that should be used to wrap the
            request context the factory should provide a render method
            returning a file like object
    """
    def _view(request):
        """
            the dynamically builded view object
        """
        filename = filename_builder(request)
        result = factory(request.context).render()
        request = write_file_to_request(request, filename, result)
        return request.response
    return _view
